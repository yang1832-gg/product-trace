#!/usr/bin/env python3
"""Build static verification data and print-ready QR codes from an Excel file."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


TOKEN_NAMESPACE = "product-trace-static-v1"
CODE_PATTERN = re.compile(r"^[0-9]{16}$")
TOKEN_PATTERN = re.compile(r"^[a-f0-9]{24}$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--site-output", type=Path)
    parser.add_argument("--skip-site-data", action="store_true")
    parser.add_argument("--mapping-output", required=True, type=Path)
    parser.add_argument("--qr-zip", type=Path)
    parser.add_argument("--base-url", required=True)
    return parser.parse_args()


def normalize_time(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    text = str(value or "").strip()
    for pattern in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    raise ValueError(f"Row {row_number}: invalid certification time: {text!r}")


def make_token(code: str, certification_time: str, batch: str) -> str:
    payload = f"{TOKEN_NAMESPACE}|{code}|{certification_time}|{batch}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def verification_url(base_url: str, token: str) -> str:
    normalized_base_url = f"{base_url.rstrip('/')}/"
    return f"{normalized_base_url}?id={token}#home"


def load_records(input_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    records: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    seen_tokens: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        code = str(row[0] or "").strip()
        certification_time = normalize_time(row[1], row_number)
        batch = str(row[2] or "").strip()

        if not CODE_PATTERN.fullmatch(code):
            raise ValueError(f"Row {row_number}: code must contain exactly 16 digits")
        if not batch:
            raise ValueError(f"Row {row_number}: product batch is empty")
        if code in seen_codes:
            raise ValueError(f"Row {row_number}: duplicate code: {code}")

        token = make_token(code, certification_time, batch)
        if not TOKEN_PATTERN.fullmatch(token) or token in seen_tokens:
            raise ValueError(f"Row {row_number}: duplicate or invalid query token")

        seen_codes.add(code)
        seen_tokens.add(token)
        records.append(
            {
                "token": token,
                "code": code,
                "time": certification_time,
                "batch": batch,
            }
        )

    if not records:
        raise ValueError("The spreadsheet contains no data rows")
    return records


def write_site_data(
    records: list[dict[str, str]], output_dir: Path, base_url: str
) -> None:
    if output_dir.exists() and any(output_dir.iterdir()):
        raise FileExistsError(f"Site output directory is not empty: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    shards: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    for record in records:
        token = record["token"]
        shards[token[:2]][token] = {
            "code": record["code"],
            "batch": record["batch"],
        }

    for shard, values in sorted(shards.items()):
        target = output_dir / f"{shard}.json"
        target.write_text(
            json.dumps(values, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

    manifest = {
        "version": 1,
        "recordCount": len(records),
        "shardCount": len(shards),
        "baseUrl": f"{base_url.rstrip('/')}/",
    }
    (output_dir.parent / "record-manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def write_mapping(
    records: list[dict[str, str]], output_path: Path, base_url: str
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "\u5e8f\u53f7",
                "\u9632\u4f2a\u7f16\u7801",
                "\u521d\u6b21\u8ba4\u8bc1\u65f6\u95f4",
                "\u4ea7\u54c1\u6279\u53f7",
                "\u67e5\u8be2ID",
                "\u4e8c\u7ef4\u7801\u7f51\u5740",
                "\u4e8c\u7ef4\u7801\u6587\u4ef6",
            ]
        )
        for index, record in enumerate(records, start=1):
            writer.writerow(
                [
                    index,
                    record["code"],
                    record["time"],
                    record["batch"],
                    record["token"],
                    verification_url(base_url, record["token"]),
                    f"{index:05d}.png",
                ]
            )


def write_qr_zip(
    records: list[dict[str, str]], output_path: Path, base_url: str
) -> None:
    try:
        import qrcode
        from qrcode.constants import ERROR_CORRECT_M
    except ImportError as error:
        raise RuntimeError(
            "QR generation requires the qrcode package: pip install qrcode[pil]"
        ) from error

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, record in enumerate(records, start=1):
            qr = qrcode.QRCode(
                version=None,
                error_correction=ERROR_CORRECT_M,
                box_size=12,
                border=4,
            )
            qr.add_data(verification_url(base_url, record["token"]))
            qr.make(fit=True)
            image = qr.make_image(fill_color="black", back_color="white")
            buffer = io.BytesIO()
            image.save(buffer, format="PNG")
            archive.writestr(f"qr/{index:05d}.png", buffer.getvalue())

            if index % 500 == 0 or index == len(records):
                print(f"Generated QR codes: {index}/{len(records)}", flush=True)

        archive.writestr(
            "README.txt",
            (
                f"Base URL: {base_url.rstrip('/')}\n"
                f"QR count: {len(records)}\n"
                "Use the external CSV mapping to match each numbered PNG to a product.\n"
            ),
        )


def main() -> int:
    args = parse_args()
    if not args.skip_site_data and args.site_output is None:
        raise ValueError("--site-output is required unless --skip-site-data is used")
    records = load_records(args.input)
    if not args.skip_site_data:
        write_site_data(records, args.site_output, args.base_url)
    write_mapping(records, args.mapping_output, args.base_url)
    if args.qr_zip:
        write_qr_zip(records, args.qr_zip, args.base_url)
    print(
        json.dumps(
            {
                "records": len(records),
                "siteOutput": str(args.site_output) if args.site_output else None,
                "mappingOutput": str(args.mapping_output),
                "qrZip": str(args.qr_zip) if args.qr_zip else None,
                "baseUrl": args.base_url.rstrip("/"),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        raise SystemExit(1)
